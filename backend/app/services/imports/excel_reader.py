from pathlib import Path

from openpyxl import load_workbook


class ExcelReader:
    """
    Utility untuk membaca file Excel (.xlsx).
    """

    def read(self, file_path: str | Path) -> tuple[list[str], list[tuple]]:
        """
        Membaca header dan seluruh data.

        Returns
        -------
        tuple
            (
                headers,
                rows,
            )
        """

        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True,
        )

        worksheet = workbook.active

        rows = list(
            worksheet.iter_rows(
                values_only=True,
            )
        )

        workbook.close()

        if not rows:
            return [], []

        headers = [
            str(value).strip()
            if value is not None
            else ""
            for value in rows[0]
        ]

        return headers, rows[1:]